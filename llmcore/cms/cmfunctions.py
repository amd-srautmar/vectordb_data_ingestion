import pandas as pd
from openpyxl import load_workbook
from bs4 import BeautifulSoup, NavigableString
import tiktoken
import weaviate  # NOT USED
import requests
import warnings  # NOT USED
import openai  # NOT USED
import time
import tenacity

# import config
import backoff

"""MOVE this section
# Configuration
openai.api_type = "azure"
openai.api_key = config.API_KEY
openai.api_base = config.RESOURCE_ENDPOINT
openai.api_version = "2023-05-15"

SERVER = config.API_PROXY_HOST
HEADERS = {"Ocp-Apim-Subscription-Key": config.API_PROXY_KEY}

if config.VECTORDB_API_KEY is not None:
    weaviate_resource_config = weaviate.AuthApiKey(
        api_key=config.VECTORDB_API_KEY
        )
    client = weaviate.Client(
        url=config.VECTORDB_HOST,
        auth_client_secret=weaviate_resource_config
        )
else:
    client = weaviate.Client(url=config.VECTORDB_HOST)
"""

# -------------------------------------------------------------------------
# LLM API Gateway Functions
# -------------------------------------------------------------------------


# From llmcore-chatbot
def api_call(
    endpoint_name: str,
    deployment_id: str,
    body: dict,
    api_proxy_url: str,
    api_proxy_key: str,
):

    headers = {"Ocp-Apim-Subscription-Key": api_proxy_key}
    response = requests.post(
        url=f"{api_proxy_url}/azure/engines/{deployment_id}/{endpoint_name}",
        json=body,
        headers=headers,
    )

    response.raise_for_status()
    return response.json()


def get_embedding_llm(input, engine, api_proxy_url, api_proxy_key):
    body = {"input": input, "user": None}
    embeddings_result = api_call(
        endpoint_name="embeddings",
        deployment_id=engine,
        body=body,
        api_proxy_url=api_proxy_url,
        api_proxy_key=api_proxy_key,
    )
    return embeddings_result["data"][0]["embedding"]


# -------------------------------------------------------------------------
# Weaviate Functions
# -------------------------------------------------------------------------


# TODO: discuss converging on this function "create_class"
def create_class(class_name, properties, client, verbose=False):
    if verbose:
        print(f"Creating class {class_name}...")

    class_obj = {"class": class_name, "properties": properties}

    if client.schema.exists(class_name=class_name):
        print(f"Class {class_name} exists already; nothing done...")
        return

    return client.schema.create_class(class_obj)


# TODO: discuss converging on this function "create_class"
def _create_class(class_name, properties, client):
    print(f"Creating class {class_name}...")
    class_obj = {"class": class_name, "properties": properties}

    new_class = client.schema.create_class(class_obj)
    return new_class


# TODO: discuss converging on this function "create_objects"
@backoff.on_exception(backoff.expo, tenacity.RetryError)
def create_objects(
    vector_class_config, df, client, batch_size=100, dynamic=False, verbose=False
):

    if verbose:
        print(f"Creating objects for class {vector_class_config['class']}...")

    client.batch.configure(batch_size=batch_size, dynamic=dynamic)

    counter = 0
    interval = 10

    for _, row in df.iterrows():
        properties = vector_class_config["properties"]
        if "properties_map" in vector_class_config:
            properties = _compose_properties_content(
                row_series=row, properties=vector_class_config["properties_map"]
            )

        vector = row["embedding_using_ada"]
        client.batch.add_data_object(
            data_object=properties,
            class_name=vector_class_config["class"],
            vector=vector,
        )

        counter += 1
        if counter % interval == 0:
            if verbose:
                print(f"Imported {counter} articles...")

    client.batch.flush()

    if verbose:
        print(f"Finished importing {counter} articles.")


# TODO: discuss converging on this function "create_objects"
def create_objects(class_name, df, client):
    print(f"Creating objects for class {class_name}...")
    # Configure the batch size
    client.batch.configure(batch_size=100)

    # Settings for displaying the import progress
    counter = 0
    interval = 100  # print progress every this many records

    # Iterate through the DataFrame rows and add each record to the batch
    for index, row in df.iterrows():

        properties = {
            "text": row["Extracted Text"],
            "n_tokens": row["n_tokens"],
            "url": row["Url"],
        }

        # Convert the vector from string back to array of floats
        vector = row["embeddeding_using_ada"]
        # Print properties and vector for debugging
        # print(f"Properties: {properties}")
        # print(f"Index: {index}")
        # Add the object to the batch, and set its vector embedding
        client.batch.add_data_object(properties, class_name, vector=vector)

        # Calculate and display progress
        counter += 1
        if counter % interval == 0:
            print(f"Imported {counter} articles...")

    client.batch.flush()
    print(f"Finished importing {counter} articles.")


def data_load_weaviate(
    class_name,
    df_to_load,
    weaviate_client,
    column_dict,
    batch_size,
    weaviate_api_key=None,
):
    """Function to load the vector data into Weaviate.

    Args:
        class_name (str): Weaviate class name
        properties (list of dict): Weaviate class properties
        df_to_load (DataFrame): Dataframe to load in Weaviate
        weaviate_client (str): Weaviate client url
        column_dict (dict): Dataframe columns passed in form of dictionary.
        batch_size (int): batch size
        weaviate_api_key (str, optional): weaviate api key for client authentication. Defaults to None.
    """
    # Create the Weaviate class
    try:
        if weaviate_api_key:
            weaviate_resource_config = weaviate.AuthApiKey(api_key=weaviate_api_key)
            client = weaviate.Client(
                url=weaviate_client, auth_client_secret=weaviate_resource_config
            )
        else:
            client = weaviate.Client(
                url=weaviate_client
            )
    except Exception as e:
        print(f"Failed to connect to Weaviate with error {e}")

    client.batch.configure(batch_size=int(batch_size))

    with client.batch as batch:
        for index, row in df_to_load.iterrows():
            article_object = {}
            for col_key in column_dict.keys():
                article_object[column_dict[col_key]] = row[column_dict[col_key]]

            batch.add_data_object(
                data_object=article_object,
                class_name=class_name,
                uuid=row["chunk_id"],
                vector=row["vector"],
            )


def delete_filedata_from_vectordb(
    class_name, weaviate_client, file_list_df, identifier: str, weaviate_api_key=None
):
    """Function to delete data from Weaviate vector db.

    Args:
        class_name (str): Weaviate class name
        weaviate_client (str): Weaviate client url
        file_list_df (DataFrame): Input dataframe
        identifier (str): identifier for deleting from Weaviate
        weaviate_api_key (str, optional): weaviate api key for client authentication. Defaults to None.
    """
    try:
        if weaviate_api_key:
            weaviate_resource_config = weaviate.AuthApiKey(api_key=weaviate_api_key)
            client = weaviate.Client(
                url=weaviate_client, auth_client_secret=weaviate_resource_config
            )
        else:
            client = weaviate.Client(
                url=weaviate_client
            )

    except Exception as e:
        print(f"Failed to connect to Weaviate with error {e}")

    for id_val in file_list_df[identifier]:
        try:
            client.batch.delete_objects(
                class_name=class_name,
                where={"path": [identifier], "operator": "Equal", "valueText": id_val},
            )
            print(f"Data Deleted from Weaviate for the file: {id_val}")

        except Exception as e:
            print(f"Exception occured while deleting data for {id_val}, error: ", e)


# -------------------------------------------------------------------------
# Soup Functions
# -------------------------------------------------------------------------


def process_tag(tag):
    if tag.name == "a":
        href = tag.get("href", "")
        if href.startswith("http"):
            return f"{tag.text} ({href})"
        else:
            return f"{tag.text} (https://amd.service-now.com/{href})"
    else:
        return tag.text


def process_table(table):
    table_data = ["Table: "]
    for row in table.find_all("tr"):
        row_data = []
        for cell in row.find_all(["td", "th"]):
            row_data.append(cell.get_text(strip=True))
        table_data.append(" | ".join(row_data))
    table_data.append(" Table ends here")
    return "\n".join(table_data)


def process_descendants(tag, text_parts):
    for child in tag.children:
        if isinstance(child, NavigableString):
            parent_tag = child.parent.name if child.parent else None
            if (
                parent_tag != "table"
                and parent_tag != "td"
                and parent_tag != "th"
                and parent_tag != "tr"
            ):
                text_parts.append(child.strip())
        elif child.name == "a":
            text_parts.append(process_tag(child))
        elif child.name == "table":
            text_parts.append(process_table(child))
            child.decompose()  # Decompose the table tag after processing it
        else:
            process_descendants(child, text_parts)


def process_soup(soup):
    text_parts = []

    # Remove the style tag if it exists
    style_tag = soup.find("style")
    if style_tag:
        style_tag.decompose()

    process_descendants(soup, text_parts)

    return " ".join(text_parts).strip()


# -------------------------------------------------------------------------
# KB Functions
# -------------------------------------------------------------------------


def split_kbs(row, token_limit=6000):
    text = row["Extracted Text"]
    tokens = row["n_tokens"]
    url = row["Url"]
    new_rows = []

    if tokens > token_limit:
        words = text.split()
        temp = text.split("Full Article:")[0]
        num_splits = tokens // token_limit + (1 if tokens % token_limit > 0 else 0)
        words_per_split = len(words) // num_splits

        for i in range(num_splits):
            start = i * words_per_split
            end = (i + 1) * words_per_split if i < num_splits - 1 else len(words)
            split_text = " ".join(words[start:end])
            if (
                i > 0
            ):  # Add "Full Article:" to the beginning of each split except the first one
                split_text = temp + " Part Article: " + split_text
            new_rows.append(
                {"Extracted Text": split_text, "n_tokens": end - start, "Url": url}
            )
    else:
        new_rows.append({"Extracted Text": text, "n_tokens": tokens, "Url": url})

    return new_rows


def process_kb_articles(path_to_kb_extract):
    print("Processing KB Articles")
    df_csv = pd.read_csv(path_to_kb_extract, encoding="latin1")

    # Create an empty pandas dataframe to store the extracted text
    data = {"Extracted Text": [], "Url": []}
    df = pd.DataFrame(data)
    count_empty_str = 1
    # Loop through each row in the worksheet
    for index, row in df_csv.iterrows():
        # Get the HTML content from the eighth column
        article_body_content = str(row[4])
        issue_content = str(row[5]) if row[5] is not None and row[5] != "nan" else ""
        resolution_content = (
            str(row[6]) if row[6] is not None and row[6] != "nan" else ""
        )
        question_content = str(row[7]) if row[7] is not None and row[7] != "nan" else ""
        instruction_content = (
            str(row[8]) if row[8] is not None and row[8] != "nan" else ""
        )

        article_soup = BeautifulSoup(article_body_content, "html.parser")
        article_content = process_soup(article_soup)
        if issue_content is not None and issue_content != "nan":
            issue_soup = BeautifulSoup(issue_content, "html.parser")
            issue_content = process_soup(issue_soup)
            article_content = article_content + " Related Issue: " + issue_content
        if resolution_content is not None and resolution_content != "nan":
            resolution_soup = BeautifulSoup(resolution_content, "html.parser")
            resolution_content = process_soup(resolution_soup)
            article_content = article_content + " Resolution: " + resolution_content
        if question_content is not None and question_content != "nan":
            question_soup = BeautifulSoup(question_content, "html.parser")
            question_content = process_soup(question_soup)
            article_content = article_content + " Question: " + question_content
        if instruction_content is not None and instruction_content != "nan":
            instruction_soup = BeautifulSoup(instruction_content, "html.parser")
            instruction_content = process_soup(instruction_soup)
            article_content = article_content + " Instruction: " + instruction_content
        # Ignore text that is less than 1 word (empty strings)
        if len(article_content.split()) == 0:
            if row[7] is not None and row[7] != "nan":
                article_body_content = str(row[7])
            elif row[8] is not None and row[8] != "nan":
                article_body_content = str(row[8])
            else:
                count_empty_str += 1
                print("Empty String: " + str(row[1]))
                continue
            article_soup = BeautifulSoup(article_body_content, "html.parser")
            article_content = process_soup(article_soup)
        if str(row[2]).strip() != "" and str(row[2]).strip() != "nan":
            article_content = article_content + " Category: " + str(row[2])
        if str(row[9]).strip() != "" and str(row[9]).strip() != "nan":
            article_content = article_content + " Metadata: " + str(row[9])

        text_list = [article_content]
        text_list = [
            "Short Description : " + row[0] + " Full Article: " + s for s in text_list
        ]
        url_list = [
            "https://amd.service-now.com/sp?id=kb_article_view&sysparm_article="
            + row[1]
        ]

        # Append the extracted text to the pandas dataframe
        for i in range(len(text_list)):
            df = df.append(
                {"Extracted Text": text_list[i], "Url": url_list[i]}, ignore_index=True
            )
    tokenizer = tiktoken.get_encoding("cl100k_base")
    df["n_tokens"] = df["Extracted Text"].apply(lambda x: len(tokenizer.encode(x)))
    # Process the DataFrame to split rows with more than 4000 tokens
    new_data = []
    for _, row in df.iterrows():
        new_rows = split_kbs(row, token_limit=config.MAX_TOKENS)
        new_data.extend(new_rows)

    # Create a new DataFrame with the split text
    new_df = pd.DataFrame(new_data)
    return new_df


# -------------------------------------------------------------------------
# Service Catalog Functions
# -------------------------------------------------------------------------


def process_service_catalog(path_to_service_catalog):
    print("Processing Service Catalog")
    workbook = load_workbook(path_to_service_catalog)
    worksheet = workbook.active
    # Create an empty pandas dataframe to store the extracted text
    data = {"Extracted Text": [], "Url": []}
    df = pd.DataFrame(data)
    # Loop through each row in the worksheet
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Get the HTML content from the eighth column
        description = row[2]
        name_content = "\n\n**Title**: " + str(row[0])
        short_description_content = str(row[1])
        meta_content = str(row[3])
        url = str(row[4])

        if short_description_content.strip() != "":
            name_content = (
                name_content + "\n\n**Short Description**: " + short_description_content
            )

        if str(description).strip() != "":
            description_soup = BeautifulSoup(description, "html.parser")
            description_content = process_soup(description_soup)
            name_content = (
                name_content + "\n\n**Long Description**: " + description_content
            )
        if meta_content.strip() != "":
            name_content = name_content + "\n\n**Metadata**: " + meta_content

        text_list = [name_content]
        url_list = [url]

        # Append the extracted text to the pandas dataframe
        for i in range(len(text_list)):
            df = df.append(
                {"Extracted Text": text_list[i], "Url": url_list[i]}, ignore_index=True
            )
    tokenizer = tiktoken.get_encoding("cl100k_base")
    df["n_tokens"] = df["Extracted Text"].apply(lambda x: len(tokenizer.encode(x)))
    return df


@backoff.on_exception(
    backoff.expo,
    (requests.exceptions.RequestException, requests.exceptions.ConnectionError),
)
def get_embedding(text, llm_service_config):
    server = "/".join([llm_service_config.api_proxy_url, "azure"])
    headers = {"Ocp-Apim-Subscription-Key": llm_service_config.api_proxy_key}
    embedding_engine = llm_service_config.embedding_engine

    text = text.replace("\n", " ")
    body = {"input": text, "user": None}

    response = requests.post(
        url=f"{server}/engines/{embedding_engine}/embeddings",
        json=body,
        headers=headers,
        verify=False,
    )

    if response.status_code >= 400:
        title = " ".join(
            text.split(":")[1].split(" ")[  # title after first colon
                :-2
            ]  # last two elementgs are "Page" & "content"
        ).strip()
        content = " ".join(text.split(":")[2]).strip()  # title after second colon
        print(f"{time.time()} {response.status_code}: {title}: {content[:20]}")

    response.raise_for_status()

    return response.json()["data"][0]["embedding"]
